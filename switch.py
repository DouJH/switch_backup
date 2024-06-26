import logging
import shutil
import telnetlib
import time
import concurrent.futures
from openpyxl import load_workbook
import os
from datetime import datetime
import json
import socket

logging.basicConfig(filename="backup_errors.log", level=logging.ERROR,
                    format="%(asctime)s - %(levelname)s - %(message)s")


class TelnetBackup():
    def __init__(self, host, username, password):
        self.host = host
        self.username = username
        self.password = password
        self.tn = telnetlib.Telnet()

    def login(self):
        try:
            self.tn.open(self.host, port=23)
        except Exception as e:
            logging.warning(f"Network connection failed for {self.host}: {e}")
            return False

        self.tn.read_until(b"Username: ", timeout=10)
        self.tn.write(self.username.encode('ascii') + b"\n")
        self.tn.read_until(b"Password: ", timeout=10)
        self.tn.write(self.password.encode('ascii') + b"\n")
        if self.tn.read_until(b'Please press ENTER.', timeout=3):
            self.tn.write(b'\n')
        time.sleep(2)
        return True

    def backup_h3c(self, tftp_server, region, ip):
        try:
            folder_path = os.path.join(base_folder, region, datetime.today().strftime('%Y%m%d'))
            filename = f"{self.host}-{datetime.today().strftime('%Y%m%d')}.bak1.cfg"

            backup_command = f"backup startup-configuration to {tftp_server} {filename}"
            self.tn.write(backup_command.encode('ascii') + b"\n")
            time.sleep(5)
            command_result = self.tn.read_very_eager().decode('ascii')
            source_path = filename
            destination_path = os.path.join(folder_path, filename)

            os.makedirs(folder_path, exist_ok=True)  # 确保父文件夹存在
            shutil.move(source_path, destination_path)
            print(f"H3C备份完成: {ip}")
        except Exception as e:
            print(f"H3C备份失败: {ip}")
            error_message = f"Error during backup_h3c for {self.host}: {e}"
            logging.error(error_message)

    def backup_cisco(self, tftp_server, region, password, ip):
        try:
            folder_path = os.path.join(base_folder, region, datetime.today().strftime('%Y%m%d'))
            filename = f"{self.host}-{datetime.today().strftime('%Y%m%d')}.cfg"

            self.tn.write('enable'.encode('ascii') + b'\n')
            self.tn.write(self.password.encode('ascii') + b'\n')
            self.tn.write('copy running-config tftp'.encode('ascii') + b'\n')
            self.tn.write(tftp_server.encode('ascii') + b'\n')
            backup_command = (filename.encode('ascii') + b'\n')

            self.tn.write(backup_command + b"\n")
            time.sleep(5)
            command_result = self.tn.read_very_eager().decode('ascii')
            source_path = filename
            destination_path = os.path.join(folder_path, filename)

            os.makedirs(folder_path, exist_ok=True)  # 确保父文件夹存在
            shutil.move(source_path, destination_path)
            print(f"CISCO备份完成: {ip}")
        except Exception as e:
            print(f"CISCO备份失败: {ip}")
            error_message = f"Error during backup_cisco for {self.host}: {e}"
            logging.error(error_message)

    def logout(self):
        self.tn.write(b"quit\n")


def select_network_interface(interfaces):
    while True:
        print("请选择一个网络接口（输入数字）：")
        for idx, interface in enumerate(interfaces, start=1):
            print(f"{idx}. {interface}")

        choice = input("请输入选择的网卡编号：")
        try:
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(interfaces):
                return interfaces[choice_idx]
            else:
                print("选择的编号无效，请重新输入。")
        except ValueError:
            print("输入的内容不是有效的数字，请重新输入。")


def backup_device(ip, username, password, device, tftp_server, region):
    telnet_backup = TelnetBackup(ip, username, password)
    if telnet_backup.login():
        if device.strip().lower() == "h3c":
            telnet_backup.backup_h3c(tftp_server, region, ip)
        elif device.strip().lower() == "cisco":
            telnet_backup.backup_cisco(tftp_server, region, password, ip)
        else:
            print(f"不支持设备型号: {device}")
        telnet_backup.logout()


if __name__ == "__main__":
    # 获取本机所有网络接口信息
    hostname = socket.gethostname()
    interfaces = socket.gethostbyname_ex(hostname)[-1]

    if len(interfaces) == 0:
        print("未找到任何网络接口信息，请检查网络配置。")
        exit()

    # 选择一个网络接口作为 tftp_server
    tftp_server = select_network_interface(interfaces)
    if not tftp_server:
        print("未能选择有效的网络接口，程序结束。")
        exit()

    # 读取config.json文件
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)

    # 获取base_folder的值
    base_folder = config['base_folder']  # 将单反斜杠替换为双反斜杠

    # Excel文件的路径（这里假设它在本地，不是网络路径）
    wb = load_workbook('switch.xlsx')
    sheet = wb.active

    # 创建 ThreadPoolExecutor 来执行多线程任务
    max_threads = 10  # 设置最大线程数量
    with concurrent.futures.ThreadPoolExecutor(max_threads) as executor:
        futures = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            region = row[0]
            device = row[1]
            ip = row[2]
            username = row[3]
            password = row[4]
            futures.append(executor.submit(backup_device, ip, username, password, device, tftp_server, region))

        # 等待所有任务完成
        concurrent.futures.wait(futures)

    print("备份完成")
    input()
